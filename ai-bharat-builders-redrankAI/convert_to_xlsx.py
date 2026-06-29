"""
convert_to_xlsx.py — Convert submission.csv to a formatted Excel workbook.

Produces an .xlsx file with:
  - Bold header row
  - Alternating row background colours
  - Auto-fitted column widths

Usage:
    python convert_to_xlsx.py --input ./submission.csv --output ./submission.xlsx
"""

import argparse
import csv
import logging
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

# Alternating row colours
COLOUR_ODD = "FFFFFF"   # white
COLOUR_EVEN = "EBF3FB"  # light blue
COLOUR_HEADER = "1F4E79"  # dark navy
FONT_HEADER = "FFFFFF"


def _auto_fit_columns(ws, data: list):
    """Set column widths based on max content length in each column."""
    col_widths = {}
    for row in data:
        for col_idx, cell_val in enumerate(row, 1):
            col_letter = get_column_letter(col_idx)
            text_len = len(str(cell_val)) if cell_val is not None else 0
            col_widths[col_letter] = max(col_widths.get(col_letter, 10), min(text_len + 4, 80))
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width


def main():
    """Entry point: read CSV and write formatted XLSX."""
    parser = argparse.ArgumentParser(description="Convert submission.csv to .xlsx")
    parser.add_argument("--input", default="./submission.csv", help="Input CSV path")
    parser.add_argument("--output", default="./submission.xlsx", help="Output XLSX path")
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)

    if not input_path.exists():
        log.error("Input file not found: %s", input_path)
        sys.exit(1)

    # Read CSV
    with open(input_path, "r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        rows = list(reader)

    if not rows:
        log.error("Input CSV is empty.")
        sys.exit(1)

    header = rows[0]
    data_rows = rows[1:]

    log.info("Read %d data rows from %s", len(data_rows), input_path)

    # Build workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Submission"

    # Header row
    header_fill = PatternFill("solid", fgColor=COLOUR_HEADER)
    header_font = Font(bold=True, color=FONT_HEADER, size=11)

    ws.append(header)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # Data rows with alternating colours
    fill_odd = PatternFill("solid", fgColor=COLOUR_ODD)
    fill_even = PatternFill("solid", fgColor=COLOUR_EVEN)

    for row_idx, row in enumerate(data_rows, 2):
        ws.append(row)
        fill = fill_odd if row_idx % 2 == 1 else fill_even
        for cell in ws[row_idx]:
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-fit columns
    _auto_fit_columns(ws, [header] + data_rows)

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    log.info("Saved %s (%d rows)", output_path, len(data_rows))


if __name__ == "__main__":
    main()
